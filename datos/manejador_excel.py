import os
from openpyxl import Workbook, load_workbook

class ManejadorExcel:
    def __init__(self):
        self.archivo = "concesionaria.xlsx"
        self.inicializar_excel()

    def inicializar_excel(self):
        """Crea el archivo y los encabezados, agregando la columna oculta 'Estado'"""
        if not os.path.exists(self.archivo):
            wb = Workbook()
            ws = wb.active
            # Se agregó la columna 'Estado' al final
            ws.append(["Cliente", "Teléfono", "Correo", "Marca", "Modelo", "Precio", "Fecha", "Estado"])
            wb.save(self.archivo)
            wb.close()

    def registrar_venta(self, datos_lista):
        """Registra la fila con el estado 'Activo' por defecto"""
        self.inicializar_excel()
        wb = load_workbook(self.archivo)
        ws = wb.active
        
        # Agregamos la etiqueta de "Activo" a la lista de datos antes de guardar
        datos_guardar = list(datos_lista)
        datos_guardar.append("Activo")
        
        ws.append(datos_guardar)
        wb.save(self.archivo)
        wb.close()

    def obtener_ventas(self):
        """Lee solo los datos que están Activos para mostrarlos en la tabla"""
        self.inicializar_excel()
        wb = load_workbook(self.archivo)
        ws = wb.active
        datos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Revisa que la fila tenga datos y que en la columna 8 (índice 7) diga "Activo"
            if row[0] and len(row) >= 8 and row[7] == "Activo":
                # Retornamos solo los primeros 7 datos (sin el estado) para que tu tabla no se descuadre
                datos.append(row[:7])
        wb.close()
        return datos

    def eliminar_venta(self, identificador):
        """Baja lógica: no borra la fila, solo la marca como 'Inactiva'"""
        self.inicializar_excel()
        wb = load_workbook(self.archivo)
        ws = wb.active

        # Recorremos las filas para buscar el registro
        for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Comparamos la primera columna (el Cliente) con el identificador que llega de la tabla
            if str(row[0]) == str(identificador):
                # Cambiamos el valor de la columna 8 (Estado) a "Inactivo"
                ws.cell(row=index, column=8).value = "Inactivo"
                break # Rompe el ciclo porque ya lo encontró y lo deshabilitó

        # Guardamos los cambios inmediatamente
        wb.save(self.archivo)
        wb.close()
        return True